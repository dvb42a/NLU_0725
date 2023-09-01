import shutil, logging, config
from django.views.decorators.clickjacking import xframe_options_exempt
from django.views.generic.base import View
from django.shortcuts import redirect, reverse, render
from django.http import HttpResponse, StreamingHttpResponse, JsonResponse
from django.utils import timezone
from .tools import *
from rest_framework.decorators import api_view
from Accounts.required import login_required, get_user_info
from Apps.models import Apps
from Client.models import ClientPlan
from BColor import SystemStatus
from django.db.models import Count
import os
import json
from django.shortcuts import render
from collections import Counter
import re
from datetime import datetime, timedelta
from django.urls import reverse
import openpyxl
import pandas as pd
from io import BytesIO
from urllib.parse import quote
from datetime import timedelta as AddDate
from django.db import models
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

logging.basicConfig(filename='log.txt', filemode='a')

LUIS_KEY = "b25e3e3c55204401bca92434148d298a"  # kingly.azure@gmail.com KinglyNLU金鑰
# KINGlY_MAIN_APP_ID = "20201119T161907_5ut8Bd301QC1"  # 閒聊綁定專案id
num_progress = 0


def sendconfig(Dict, request):
    try:
        print(request.session['train_name'])
        app = Apps.objects.filter(app_name=request.session['train_name']).first()
        temp = {'title': config.title,
                'description': config.description,
                'in_training': app.training,
                'in_train_name': request.session['train_name']
                }
        print(temp)
    except KeyError:
        print("no train_name")
        temp = {'title': config.title,
                'description': config.description,
                'in_training': 0
                }
    Dict.update(temp)
    return Dict


class Dashboard(View):
    @staticmethod
    @login_required
    def get(request, context):
        """
        儀表板
        """
        # 當權限為 系統管理員/專案管理員 時可以查看 CLU 任務型資料
        account = context["account"]
        client_plan = ClientPlan.objects.filter(ac=account.ac_name, plan_end__gt=timezone.now()) # 查詢登入帳號所持有還未到期的plan
        client_plan_withCountApp = client_plan.annotate(
            current_app=Count('plan_app', filter=models.Q(plan_app__state=1))
        )
        apps = Apps.objects.filter(ac=account.ac_name, state=1).order_by('-created_date')  # 查詢登入帳號所持有的app清單

        context.update(dict(
            role=account.role,
            client_plan= client_plan_withCountApp,
            apps=apps,
            success="pass",
            dashborad=True
        ))
        print(SystemStatus.INFO, f"context：{context}")
        return render(request, 'dashboard.html', sendconfig(context, request))


@login_required
def app_info_view(request, context):
    """
    任務-App資訊頁
    context內的值從required繼承而來
    """
    app_name = request.GET.get('app')
    user_id = request.session['login_name']
    app = Apps.objects.filter(app_name=app_name, ac=user_id).first()
    if app.deployed == 1:
        host_name = request.get_host()
        endpoint_url = host_name + f"/api/qa/?id={user_id}-{app_name}&q=問題"
        chat_bubble = f'''<iframe src="https://{host_name}/app/kingly/chatbox/?id={user_id}-{app_name}" width="350px" height="100%" 
            frameborder="0" allowtransparency="true" scrolling="no" allowfullscreen="true" 
            style="z-index: 9999; bottom: 23px; right: 28px; position: fixed;"></iframe>'''
    else:
        endpoint_url = ""
        chat_bubble = ""
    context.update(dict(
        app=app,
        url=f"app={app_name}",
        endpoint_url=endpoint_url,
        last_deployed_date=app.last_deployed_date,
        trained=app.trained,
        deployed=app.deployed,
        chat_bubble=chat_bubble))
    return render(request, 'app_info.html', sendconfig(context, request))

def getAllValues(start,end,count,data):

    if count is not None  and start is None :
        current_date = datetime.now().date()
        end_date = current_date
        start_date = current_date - timedelta(days=count)
        all_values = [item['clu_intent'] for item in data if item['clu_intent'] is not None
                      and start_date <= datetime.strptime(item['ask_time'],'%Y-%m-%d %H:%M:%S').date() <= end_date]
    if start is not None and count is None:
        start_date_str = start
        end_date_str = end
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        all_values = [item['clu_intent'] for item in data if item['clu_intent'] is not None
                      and start_date <= datetime.strptime(item['ask_time'],
                                                          '%Y-%m-%d %H:%M:%S').date() <= end_date]
    if start is None and count is None :
        all_values = [item['clu_intent'] for item in data if item['clu_intent'] is not None]

    return all_values

def getAllValuesNone(start,end,count,data,search):
    if not search:
        search_term=""
    else:
        search_term = search  # Get the search term from the request
    pattern = re.compile(search_term, re.IGNORECASE)

    if count is not None  and start is None:
        print('123')
        current_date = datetime.now().date()
        end_date = current_date
        start_date = current_date - timedelta(days=count)
        all_values = [item for item in data if
                              item['clu_intent'] == "None" and
                              (re.search(pattern, item['q']) or
                               any(re.search(pattern, e['entity']) for e in item['entities']) or
                               any(re.search(pattern, e['entity_category']) for e in item['entities']))
                      and start_date <= datetime.strptime(item['ask_time'],'%Y-%m-%d %H:%M:%S').date() <= end_date]
    if start is not None and count is None:
        print('456')
        start_date_str = start
        end_date_str = end
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        all_values = [item for item in data if
                              item['clu_intent'] == "None" and
                              (re.search(pattern, item['q']) or
                               any(re.search(pattern, e['entity']) for e in item['entities']) or
                               any(re.search(pattern, e['entity_category']) for e in item['entities']))
                      and start_date <= datetime.strptime(item['ask_time'],
                                                          '%Y-%m-%d %H:%M:%S').date() <= end_date]
    if start is None and count is None:
        print('789')
        all_values = [item for item in data if
                      item['clu_intent'] == "None" and
                      (re.search(pattern, item['q']) or
                       any(re.search(pattern, e['entity']) for e in item['entities']) or
                       any(re.search(pattern, e['entity_category']) for e in item['entities']))]

    return all_values


@login_required
def app_result_view(request, context):
    """
    任務-App資訊頁
    context內的值從required繼承而來
    """
    app_name = request.GET.get('app')
    user_id = request.session['login_name']
    app = Apps.objects.filter(app_name=app_name, ac=user_id).first()
    app_datetime=app.created_date.strftime('%y%m%d%H%M')
    dataFileName=user_id+'-'+app_name+'-'+app_datetime
    project_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # 构建 JSON 文件的完整路径
    json_file_path = os.path.join(project_path, 'QAHistory/'+dataFileName+'.json')

    try:
        # 打开并读取 JSON 文件
        with open(json_file_path, 'r', encoding='utf-8') as json_file:

            json_data = json.load(json_file)
            item_count = len(json_data)

            if "date" in request.GET:
                current_date = datetime.now().date()
                start = None
                end = None
                date = request.GET['date']
                if date == "half":
                    countDate = 15
                if date == "1":
                    countDate = 30
                if date == "3":
                    countDate = 90
                if date == "6":
                    countDate = 180
            if "start" in request.GET:
                start = request.GET['start']
                end = request.GET['end']
                countDate = None
            if "start" not in request.GET and "date" not in request.GET:
                start=None
                end=None
                countDate=None

            print(start)
            print(end)
            print(countDate)
            all_values = getAllValues(start, end, countDate, json_data)

            value_counter = Counter(all_values)
            result = [{"name": value, "count": count} for value, count in value_counter.items()]
            sorted_result=sorted(result , key=lambda x: x['count'], reverse=True)
            context['results']=sorted_result
    except FileNotFoundError:
        print('Error!!!')

    context.update(dict(
        app=app,
        url= f"app={app_name}",
        last_deployed_date=app.last_deployed_date,
        trained=app.trained,
        deployed=app.deployed,))
    return render(request, 'app_result.html', sendconfig(context, request))

@login_required
def app_none_view(request, context):
    """
    任務-App資訊頁
    context內的值從required繼承而來
    """
    app_name = request.GET.get('app')
    user_id = request.session['login_name']
    app = Apps.objects.filter(app_name=app_name, ac=user_id).first()
    datetimeApp=app.created_date.strftime('%y%m%d%H%M')
    dataFileName=user_id+'-'+app_name+'-'+datetimeApp
    project_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # 构建 JSON 文件的完整路径
    json_file_path = os.path.join(project_path, 'QAHistory/'+dataFileName+'.json')

    try:
        # 打开并读取 JSON 文件
        with open(json_file_path, 'r', encoding='utf-8') as json_file:

            json_data = json.load(json_file)
            item_count = len(json_data)
            search = request.GET.get('search')

            if "date" in request.GET:
                current_date = datetime.now().date()
                start=None
                end=None
                date = request.GET['date']
                if date == "half":
                    countDate = 15
                if date == "1":
                    countDate = 30
                if date == "3":
                    countDate = 90
                if date == "6":
                    countDate = 180
            if "start" in request.GET:
                start= request.GET['start']
                end = request.GET['end']
                countDate = None
            if "start" not in request.GET and "date" not in request.GET:
                start= None
                end = None
                countDate = None

            all_values=getAllValuesNone(start,end,countDate,json_data,search)
            context['results']=all_values
    except FileNotFoundError:
        print('445')

    context.update(dict(
        app=app,
        url=f"app={app_name}",
        last_deployed_date=app.last_deployed_date,
        trained=app.trained,
        deployed=app.deployed,))
    return render(request, 'app_noneList.html', sendconfig(context, request))

@login_required
def app_excel_result(request, context):
    app_name = request.GET.get('app')
    user_id = request.session['login_name']
    app = Apps.objects.filter(app_name=app_name, ac=user_id).first()
    current_date = datetime.now().date().strftime('%Y%m%d')

    # Find out the json file of the app
    datetimeAPP = app.created_date.strftime('%y%m%d%H%M')
    dataFileName = user_id + '-' + app_name + '-' + datetimeAPP
    project_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    json_file_path = os.path.join(project_path, 'QAHistory/' + dataFileName + '.json')

    # Create DataFrames to hold the data
    df_result = pd.DataFrame(columns=["意圖名稱", "次數"])

    # Try to read JSON data and append to DataFrames
    try:
        with open(json_file_path, 'r', encoding='utf-8') as json_file:

            excel_name = current_date + "_" + user_id + "_" + app_name + "統計數據輸出"
            data = json.load(json_file)  # Parse JSON data into a list of dictionaries

            # Filter data based on 'clu_intent' value
            filtered_data = [item for item in data if item.get('clu_intent') == "None"]

            if "date" in request.GET:
                current_date = datetime.now().date()
                start=None
                end=None
                date = request.GET['date']
                if date == "half":
                    countDate = 15
                if date == "1":
                    countDate = 30
                if date == "3":
                    countDate = 90
                if date == "6":
                    countDate = 180
            if "start" in request.GET:
                start= request.GET['start']
                end = request.GET['end']
                countDate = None

            all_values=getAllValues(start,end,countDate,data)

            value_counter = Counter(all_values)
            result = [{"name": value, "count": count} for value, count in value_counter.items()]
            sorted_result = sorted(result, key=lambda x: x['count'], reverse=True)

            # Append filtered data to 'df_result'
            for item in sorted_result:
                df_result = df_result.append({"意圖名稱": item.get("name", ""), "次數": item.get("count", "")},
                               ignore_index=True)

            # Create a BytesIO buffer to store the Excel file
            excel_buffer = BytesIO()

            # Write DataFrames to Excel file
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df_result.to_excel(writer, sheet_name='查詢次數彙整', index=False)

                # Get the openpyxl workbook and worksheet objects
                workbook = writer.book
                worksheet_result = writer.sheets['查詢次數彙整']
                # Auto-adjust column width based on content
                for column_cells in worksheet_result.columns:
                    max_length = 0
                    column_letter = column_cells[0].column_letter
                    for cell in column_cells:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 4  # Adding a little extra width
                    worksheet_result.column_dimensions[column_letter].width = adjusted_width

            # Set the response's content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            # Encode the filename using URL encoding
            encoded_filename = quote(excel_name.encode('utf-8'))

            # Set the filename for download using URL-encoded filename
            response['Content-Disposition'] = f'attachment; filename="{encoded_filename}.xlsx"'

            # Write the content of the Excel buffer to the response
            response.write(excel_buffer.getvalue())

            return response
    except FileNotFoundError:

        print("JSON file not found.")

@login_required
def app_excel_none(request, context):
    app_name = request.GET.get('app')
    user_id = request.session['login_name']
    app = Apps.objects.filter(app_name=app_name, ac=user_id).first()
    current_date = datetime.now().date().strftime('%Y%m%d')

    # Find out the json file of the app
    datetimeAPP = app.created_date.strftime('%y%m%d%H%M')
    dataFileName = user_id + '-' + app_name + '-' + datetimeAPP
    project_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    json_file_path = os.path.join(project_path, 'QAHistory/' + dataFileName + '.json')

    # Create DataFrames to hold the data
    df_none = pd.DataFrame(columns=["用戶問題", "推測意圖", "系統預測類別"])

    # Try to read JSON data and append to DataFrames
    try:
        with open(json_file_path, 'r', encoding='utf-8') as json_file:

            excel_name = current_date + "_" + user_id + "_" + app_name + "統計數據輸出"
            data = json.load(json_file)  # Parse JSON data into a list of dictionaries

            # Filter data based on 'clu_intent' value
            filtered_data = [item for item in data if item.get('clu_intent') == "None"]

            if "date" in request.GET:
                current_date = datetime.now().date()
                start=None
                end=None
                date = request.GET['date']
                if date == "half":
                    countDate = 15
                if date == "1":
                    countDate = 30
                if date == "3":
                    countDate = 90
                if date == "6":
                    countDate = 180
            if "start" in request.GET:
                start= request.GET['start']
                end = request.GET['end']
                countDate = None
            if "start" not in request.GET and "date" not in request.GET:
                start= None
                end = None
                countDate = None

            all_values = getAllValuesNone(start,end,countDate,filtered_data)

            value_counter = Counter(all_values)
            result = [{"name": value, "count": count} for value, count in value_counter.items()]
            sorted_result = sorted(result, key=lambda x: x['count'], reverse=True)

            # Append filtered data to 'df_none'
            for item in sorted_result:
                df_none = df_none.append({"用戶問題": item.get("q", ""), "推測意圖": item.get("clu_intent", ""),
                                "系統預測類別": item.get("entities", [])},
                               ignore_index=True)


            # Create a BytesIO buffer to store the Excel file
            excel_buffer = BytesIO()

            # Write DataFrames to Excel file
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df_none.to_excel(writer, sheet_name='None回應報告', index=False)

                # Get the openpyxl workbook and worksheet objects
                workbook = writer.book
                worksheet_none = writer.sheets['None回應報告']
                # Auto-adjust column width based on content
                for column_cells in worksheet_none.columns:
                    max_length = 0
                    column_letter = column_cells[0].column_letter
                    for cell in column_cells:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 2  # Adding a little extra width
                    worksheet_none.column_dimensions[column_letter].width = adjusted_width



            # Set the response's content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            # Encode the filename using URL encoding
            encoded_filename = quote(excel_name.encode('utf-8'))

            # Set the filename for download using URL-encoded filename
            response['Content-Disposition'] = f'attachment; filename="{encoded_filename}.xlsx"'

            # Write the content of the Excel buffer to the response
            response.write(excel_buffer.getvalue())

            return response
    except FileNotFoundError:

        print("JSON file not found.")

@login_required
def app_intent(request, context):
    """
    任務-意圖設定頁
    """
    app_name = request.GET.get('app')
    user_id = request.session['login_name']
    app = Apps.objects.filter(app_name=app_name, ac=user_id).first()
    intent_list = []
    intent_json = f'./config/{user_id}_{app_name}/intent.json'
    if not os.path.isfile(intent_json):
        all_intent_count = 1
        all_count = 0
        temp = {'name': 'None', 'utterance': [], "utterance_count": 0}
        intent_list.append(temp)
    else:
        with open(f'{intent_json}', 'r', encoding='utf-8') as f:
            data = json.load(f)
        all_intent_count = len(data)
        all_count = 0
        for i in data:
            all_count = all_count + len(i["utterance"])
            i.update({"utterance_count": len(i["utterance"])})
        intent_list = data
    context.update(dict(
        app_name=app_name,
        url=f"app={app_name}",
        intent_list=intent_list,
        all_intent_count=all_intent_count,
        all_labeled_examples_count=all_count,
        trained=app.trained
    ))

    return render(request, 'intent.html', sendconfig(context, request))


@login_required
def intent_utterance(request, context):
    """
    任務-訓練語句頁
    """
    app_name = request.GET.get('app')
    app_intent_name = request.GET.get('intentName')
    user_id = request.session['login_name']
    intent_data = []
    file = f'./config/{user_id}_{app_name}/utterance.json'
    with open(file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    for i in data:
        try:
            if i['intent'] in app_intent_name:
                if i not in intent_data:
                    intent_data.append(i)
        except KeyError:
            pass
    context.update(dict(
        app_name=app_name,
        url=f"app={app_name}",
        app_intent_name=app_intent_name,
        labeled_examples=intent_data,
        labeled_examples_count=len(intent_data)
    ))
    return render(request, 'intent_utterance.html', sendconfig(context, request))


@login_required
def app_entity(request, context):
    """
    任務-實體頁
    """
    app_name = request.GET.get('app')
    user_id = request.session['login_name']
    entity_list = []
    entity_json = f'./config/{user_id}_{app_name}/entity.json'
    if not os.path.isfile(entity_json):
        pass
    else:
        with open(f'{entity_json}', 'r', encoding='utf-8') as f:
            data = json.load(f)
        for i in data:
            i.update({"utterance_count": len(i["utterance"])})
        entity_list = data
    context.update(dict(
        app_name=app_name,
        url=f"app={app_name}",
        entity_list=entity_list,
        entity_list_count=len(entity_list)
    ))
    return render(request, 'entity.html', sendconfig(context, request))


@login_required
def entity_utterance(request, context):
    """
    任務-實體相關的訓練語句清單頁
    """
    app_name = request.GET.get('app')
    app_entity_name = request.GET.get('entityName')
    user_id = request.session['login_name']
    entity_data = []
    file = f'./config/{user_id}_{app_name}/utterance.json'
    with open(file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    for i in data:
        for j in i["utterance"]:
            try:
                if j["category"] in app_entity_name:
                    if i not in entity_data:
                        entity_data.append(i)
            except TypeError:
                pass
    context.update(dict(
        app_name=app_name,
        url=f"app={app_name}",
        app_entity_name=app_entity_name,
        labeled_examples=entity_data,
        labeled_examples_count=len(entity_data)
    ))
    return render(request, 'entity_utterance.html', sendconfig(context, request))


@login_required
def ans_rules(request, context):
    """
    任務-對話規則頁
    """
    if request.method == 'GET':
        app_name = request.GET.get('app')
        user_id = request.session['login_name']
        intent_json = f'./config/{user_id}_{app_name}/intent.json'
        entity_json = f'./config/{user_id}_{app_name}/entity.json'
        ans_json = f'./config/{user_id}_{app_name}/ans.json'
        if not os.path.isfile(ans_json) or not os.path.isfile(intent_json):
            intent_list = []
            entity_list = []
            ans_list = []
            ans = False
        else:
            try:
                with open(intent_json, 'r', encoding='utf-8') as f:
                    intent_list = json.load(f)
                with open(entity_json, 'r', encoding='utf-8') as f:
                    entity_list = json.load(f)
                with open(ans_json, 'r', encoding='utf-8') as f:
                    ans_list = json.load(f)
                ans = True
            except FileNotFoundError:
                pass
        context.update(dict(
            app_name=app_name,
            url=f"app={app_name}",
            intent_list=intent_list,
            entity_list=entity_list,
            ans_list=ans_list,
            ans=ans
        ))
        return render(request, 'ans_rules.html', sendconfig(context, request))


class TestQAPage(View):
    @staticmethod
    @login_required
    def get(request, context):
        """
        問答測試頁
        """
        app_name = request.GET.get('app')
        context.update(dict(
            app_name=app_name,
            url=f"app={app_name}"
        ))
        return render(request, 'test_chat.html', sendconfig(context, request))


@login_required
def qa_log(request, context):
    """
    對話紀錄
    """
    app_name = request.GET.get('app')
    user_id = request.session['login_name']
    app = Apps.objects.filter(app_name=app_name, ac=user_id).first()
    date = app.created_date.strftime("%y%m%d%H%M")
    file_path = f'./QAHistory/{user_id}-{app_name}-{date}.json'
    print(file_path)
    input_data = ReadQAHistory(file_path)
    context.update(dict(
        app_name=app_name,
        url=f"app={app_name}",
        log_list=input_data,
    ))
    return render(request, 'qa_log.html', sendconfig(context, request))


class ChatBox(View):
    @staticmethod
    @xframe_options_exempt
    def get(request):
        """
        前端嵌入ChatBox
        """
        app_id = request.GET.get('id')
        return render(request, 'ChatBox.html', {"app_id": app_id})


def creat_app(request):
    """
    建立任務App
    """
    account = get_user_info(request)
    if request.method == 'POST':
        plan = ClientPlan.objects.get(id=request.POST['plan_id'])
        current_app=Apps.objects.filter(plan_id=request.POST['plan_id']).count()

        # apps = Apps.objects.filter(plan_id=plan.id).count()
        if plan.max_app > current_app:
            print("新建立中")
            Apps.objects.create(
                plan=plan,
                ac=account,
                app_name=request.POST['new_app_name'],
                app_desc=request.POST['app_description'])
            user_id = request.session['login_name']
            new_app_name = user_id + "-" + request.POST.get('new_app_name', None)
            app_description = request.POST.get('app_description', None)
            status = addApp(new_app_name, app_description)
            if status == 201 or 200:
                return JsonResponse({'report': 'success'})
            else:
                return JsonResponse({'report': "error"})
        else:
            return redirect(reverse("app:dashboard"))


def delete_app(request):
    """
    刪除任務App
    """
    app_name = request.GET.get('app')
    user_id = request.session['login_name']
    path = f'./config/{user_id}_{app_name}/'
    try:
        shutil.rmtree(path)
    except OSError:
        pass
    state = deleteApp(user_id, app_name)
    if state == 202:
        print(f'[DELETE app]: {user_id}-{app_name}\nresponse:{state}')
        app = Apps.objects.filter(app_name=app_name, ac=user_id).first()
        app.state=1
        app.deleted_date = datetime.now()
        app.save()
        return redirect(reverse("app:dashboard"))
    else:
        return HttpResponse('<h1>App 刪除失敗</h1><br /><a href="/app/">Back page</a><br />' + str(state))

def delete_app_in_ClientPlan(request):
    app_name = request.GET.get('app')
    user_id = request.session['login_name']
    path = f'./config/{user_id}_{app_name}/'
    try:
        shutil.rmtree(path)
    except OSError:
        pass
    state = deleteApp(user_id, app_name)
    if state == 202:
        print(f'[DELETE app]: {user_id}-{app_name}\nresponse:{state}')
        app = Apps.objects.filter(app_name=app_name, ac=user_id).first()
        app.state = 1
        app.deleted_date = datetime.now()
        app.save()
        previous_page = request.META.get('HTTP_REFERER', '/')
        return redirect(previous_page)
    else:
        return HttpResponse('<h1>App 刪除失敗</h1><br /><a href="/app/">Back page</a><br />' + str(state))

def train_app(request):
    """
    任務-訓練模型
    """
    app_name = request.GET.get('app')
    user_id = request.session['login_name']
    app = Apps.objects.filter(app_name=app_name, ac=user_id).first()
    train_name = f"train_V{app.train_version}"
    print(f"{user_id}-----{app_name}-----{train_name}")
    res = trainApp(user_id, app_name, train_name)
    if res.status_code == 202:
        app.deployed = 0
        app.training = 1
        app.save(update_fields=['deployed', 'training'])
        request.session['train_name'] = app_name
        return HttpResponse(200)
    else:
        return HttpResponse(str(res.text))


def train_event(request):
    def stream_generator():
        user_id = request.session['login_name']
        app = Apps.objects.filter(app_name=request.session['train_name'], ac=user_id).first()
        while True:
            train_dict = getTrainStatus(user_id, request.session['train_name'])
            for i in range(len(train_dict["value"])):
                if train_dict["value"][i]["status"] in "running":
                    r = i
            train_process = int(train_dict["value"][r]["result"]["trainingStatus"]["percentComplete"] * 0.8) + \
                            int(train_dict["value"][r]["result"]["evaluationStatus"]["percentComplete"] * 0.2)
            data = {'in_train_name': request.session['train_name'], 'train_process': train_process}
            yield u'data: %s\n\n' % str(json.dumps(data))
            time.sleep(3)
            if train_process == 100:
                app.last_trained_date = timezone.now()
                app.train_version = app.train_version + 1
                app.trained = 1
                app.training = 0
                app.save(update_fields=['last_trained_date', 'train_version', 'trained', 'training'])
                request.session['train_name'] = None
                break

    response = StreamingHttpResponse(stream_generator(), content_type="text/event-stream")
    response['Cache-Control'] = 'no-cache'
    return response


def update_app_description(request):
    """
    任務-更新App描述
    """
    if request.is_ajax():
        app_name = request.POST.get('app_name')
        user_id = request.session['login_name']
        put_app_description = request.POST.get('put_app_description')
        app = Apps.objects.filter(app_name=app_name, ac=user_id).first()
        res = addApp(f'{user_id}-{app_name}', put_app_description)
        if res == 200:
            app.app_desc = put_app_description
            app.save()
            return HttpResponse(200)
        else:
            return HttpResponse(str(res))


def app_publish(request):
    """
    任務-App發布
    """
    if request.is_ajax():
        app_name = request.GET.get('app')
        user_id = request.session['login_name']
        app = Apps.objects.filter(app_name=app_name, ac=user_id).first()
        train_name = f"train_V{app.train_version - 1}"
        deploy_name = f"deploy_V{app.deploy_version}"
        print(user_id, app_name, train_name, deploy_name)
        res = deployApp(user_id, app_name, train_name, deploy_name)
        print(res.status_code)
        if res.status_code == 202:
            while True:
                try:
                    publish_dict = deployInfo(user_id, app_name, deploy_name)
                    print("-----------------" + str(publish_dict))
                    publish_time_str = publish_dict['lastDeployedDateTime']
                    dt = datetime.strptime(publish_time_str, "%Y-%m-%dT%H:%M:%SZ")
                    adjusted_dt = dt + AddDate(hours=8)
                    app.last_deployed_date = adjusted_dt.strftime("%Y-%m-%d %H:%M:%S")
                    app.deployed = 1
                    app.deploy_version = app.deploy_version + 1
                    app.save(update_fields=['last_deployed_date', 'deployed', 'deploy_version'])
                    return JsonResponse({'report': 'success'})
                    break
                except KeyError:
                    time.sleep(2)
                    pass

        else:
            publish_dict = deployInfo(user_id, app_name, deploy_name)
            return JsonResponse({'error_msg': publish_dict['error']['code']}, status=400)


def upload_base2excel(request):
    """
    任務-匯入基本包
    """
    global num_progress
    num_progress = 0
    user_id = request.session['login_name']
    app_name = request.POST.get('app_name')
    # print(app_name)
    path = f'./config/{user_id}_{app_name}/'
    if not os.path.exists(path):
        os.makedirs(path)
    if request.is_ajax():
        num_progress = 50
        excel = request.FILES.get('file')
        wb = openpyxl.load_workbook("./base.xlsx")
        base_inf = openpyxl.load_workbook(excel)
        try:
            entity_sheet = wb["Entity"]
            utterance_sheet = wb["Utterance"]
            ans_sheet = wb["Ans"]
            information_sheet = base_inf["Information"]
        except KeyError:
            return JsonResponse({'message': 'error'})
        rows_inf = information_sheet.rows
        base_dict = {}
        base_list = []
        utt_list = ['#部門', '#產品']
        # 處理entity
        for row in list(rows_inf):  # 遍歷每行資料
            temp = []  # 用於存放一行資料
            for c in row:  # 把每行的每個單元格的值取出來，存放到case裡
                if str(c.value) not in "None":
                    temp.append(str(c.value))
            if len(temp) > 0:
                temp_list = []
                for i in range(len(temp) - 1):
                    temp_list.append(temp[i + 1])
                temp_dict = {temp[0]: temp_list}
                if len(temp_list) > 0:
                    base_list.append(temp[0])
            base_dict.update(temp_dict)
        num_progress = 75

        cols_ent = entity_sheet.columns
        for col in list(cols_ent):
            for c in col:
                if c.value is not None:
                    if c.value == "產品" or c.value == "部門":
                        s = "#" + c.value
                        for i in range(len(base_dict[s])):
                            target_row = c.row + i + 1
                            target_column = c.column
                            entity_sheet.cell(row=target_row, column=target_column).value = base_dict[s][i]
                    else:
                        for i in base_list:
                            if i in str(c.value):
                                c.value = str(c.value).replace(i, base_dict[i][0])

        # 處理utterance
        rows_utt = utterance_sheet.rows
        for row in list(rows_utt):
            for c in row:
                if c.value is not None:
                    for i in range(len(utt_list)):
                        for j in range(len(base_dict[utt_list[i]])):
                            if c.value == "問" + utt_list[i]:
                                intent = str(c.value).replace(utt_list[i], base_dict[utt_list[i]][j])
                                utter = str(utterance_sheet.cell(row=c.row, column=c.column + 1).value).replace(
                                    utt_list[i], base_dict[utt_list[i]][j])
                                utterance_sheet.append([intent, utter])
                    for k in base_list:
                        if k in str(c.value):
                            c.value = str(c.value).replace(k, base_dict[k][0])

        # 處理ans
        rows_ans = ans_sheet.rows
        for row in list(rows_ans):
            for c in row:
                if c.value is not None:
                    for i in range(len(utt_list)):
                        for j in range(len(base_dict[utt_list[i]])):
                            if c.value == "問" + utt_list[i]:
                                intent = str(c.value).replace(utt_list[i], base_dict[utt_list[i]][j])
                                ans1 = str(ans_sheet.cell(row=c.row, column=c.column + 1).value).replace(utt_list[i],
                                                                                                         base_dict[
                                                                                                             utt_list[
                                                                                                                 i]][j])
                                ans2 = str(ans_sheet.cell(row=c.row, column=c.column + 2).value).replace(utt_list[i],
                                                                                                         base_dict[
                                                                                                             utt_list[
                                                                                                                 i]][j])
                                for l in base_list:
                                    if l in ans1:
                                        ans1 = ans1.replace(l, base_dict[l][j])
                                    if l in ans2:
                                        ans2 = ans2.replace(l, base_dict[l][j])
                                ans_sheet.append([intent, ans1, ans2])
                for k in base_list:
                    if k in str(c.value):
                        c.value = str(c.value).replace(k, base_dict[k][0])
        utterance_sheet.delete_rows(45, 14)
        ans_sheet.delete_rows(7, 2)
        wb.save(f'./config/{user_id}_{app_name}/base.xlsx')
        num_progress = 100
        time.sleep(1.5)
        num_progress = 0
        return JsonResponse({'message': 'success'})


def check_base(request):
    """
    任務-檢查base.excel
    """
    user_id = request.session['login_name']
    app_name = request.POST.get('app_name')
    print(request)
    if os.path.isfile(f'./config/{user_id}_{app_name}/base.xlsx'):
        return JsonResponse({'message': 'exist'})
    else:
        return JsonResponse({'message': 'empty'})


def upload_excel2json(request):
    """
    任務-匯入訓練語句
    """
    global num_progress
    num_progress = 0
    user_id = request.session['login_name']
    app_name = request.POST.get('app_name')
    app = Apps.objects.filter(app_name=app_name, ac=user_id).first()
    desc = app.app_desc
    if request.is_ajax():
        excel = request.FILES.get('file')
        base = request.POST.get('base')
        num_progress = 25
        all_count = excel_to_json(excel, user_id, app_name, desc, base)
        if all_count[0] > 0 and (len(all_count[1]) != 0 or len(all_count[2]) != 0):
            return JsonResponse({'report': "miss", 'uttr_intent': all_count[1], 'ans_intent': all_count[2]})
        elif all_count[0] == -1:
            return JsonResponse({'report': "error"})
        elif all_count[0] == -2:
            return JsonResponse({'report': "same", 'same_entity': all_count[1]})
        num_progress = 50
        state = importApp(user_id, app_name)
        if state != 202:
            return JsonResponse({'report': "fail"})
        num_progress = 75
        redefine_utter(user_id, app_name)
        try:
            os.remove(f'./config/{user_id}_{app_name}/entity.json')
            os.remove(f'./config/{user_id}_{app_name}/intent.json')
            os.remove(f'./config/{user_id}_{app_name}/ans.json')
        except FileNotFoundError:
            pass
        os.rename(f'./config/{user_id}_{app_name}/temp_entity.json', f'./config/{user_id}_{app_name}/entity.json')
        os.rename(f'./config/{user_id}_{app_name}/temp_intent.json', f'./config/{user_id}_{app_name}/intent.json')
        os.rename(f'./config/{user_id}_{app_name}/temp_ans.json', f'./config/{user_id}_{app_name}/ans.json')
        num_progress = 100
        app.trained = 0
        app.save(update_fields=['trained'])
        time.sleep(1.5)
        num_progress = 0
        return JsonResponse({'app_name': app_name})


def show_progress(request):
    """
    進度條
    """
    return JsonResponse(num_progress, safe=False)


def send_q_to_bot(request):
    """
    問答測試頁-問答ajax api
    """
    if request.is_ajax():
        time_start = time.time()
        user_id = request.session['login_name']
        app_name = request.POST.get('app_name')
        app = Apps.objects.filter(app_name=app_name, ac=user_id).first()
        q = request.POST.get('q')
        q = q.strip()
        q = q.lower()
        app_id = f'{user_id}-{app_name}'
        print(app_id + '================' + q)
        deploy_name = f"deploy_V{app.deploy_version - 1}"
        # CLU API
        result = predictCLU(time_start, app_id, deploy_name, q)
        return JsonResponse({'replies_msg': result})


def check_qnot(s):
    """
    處理單、雙引號
    """
    s = s.replace("'", "%20")
    s = s.replace('"', '%21')
    return s


@api_view(['GET'])
def qa_endpoint_api(request):
    """
    App終端問答api
    """
    time_start = time.time()
    # x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    # # print(x_forwarded_for)
    # if x_forwarded_for:
    #     remote_addr = x_forwarded_for.split(',')[0]
    # else:
    #     remote_addr = request.META.get('REMOTE_ADDR')
    # print("remote ip: " + str(remote_addr))
    app_id = request.GET.get('id')
    # user = request.GET.get('user', '端點API')
    # user_info = check_user_info(user)
    app = Apps.objects.filter(app_name=app_id.split('-')[-1], ac=app_id.split('-')[0]).first()
    q = request.GET.get('q')
    q = q.replace(",", "")
    q = q.lower()
    deploy_name = f"deploy_V{app.deploy_version - 1}"
    print(app_id + '-----------' + q)
    # CLU API
    result = predictCLU(time_start, app_id, deploy_name, q)
    print(result)
    if result['ans'] in "error":
        result = {
            'q': q,
            'intent': "Err",
            'score': str(0),
            'ans': "Err",
            'entities': [],
            # 'rule_name': "Err",
            # 'user': user,
            'run_time': "0 秒",
            'ask_time': result['ask_time']
        }
    return HttpResponse(json.dumps(result, indent=3, ensure_ascii=False), content_type="application/json")


def page_not_found(request):
    return render(request, '404.html')

